import streamlit as st
import pandas as pd
import fitz
import re
from openpyxl import load_workbook
from io import BytesIO

def estrai_dati(pdf_stream, tipo):
    pdf_document = fitz.open(stream=pdf_stream, filetype="pdf")
    testo_pdf = ''
    for page_num in range(len(pdf_document)):
        page = pdf_document.load_page(page_num)
        testo_pdf += page.get_text("text")

    if tipo == "Tipo 1 (PV014-PE-...)":
        pattern = re.compile(r"(PV0\d{2}-PE-[A-Z]{4}-[A-Z]{3}-\d{5}-[A-Z]{3}-\d{6})\s+(.+?)(?=PV0\d{2}|$)", re.DOTALL)
    elif tipo == "Tipo 2 (1.1.1)":
        pattern = re.compile(r"(\d+\.\d+\.\d+)\s+(.*?)\s+\d{2}/\d{2}/\d{4}", re.DOTALL)
    elif tipo == "Tipo 3 (T00EG00...)":
        pattern = re.compile(r"(T00EG\d{2}[A-Z]{3,6}\d{2}[A-Z]?)\s+(.*?)\s+A\d", re.DOTALL)
    elif tipo == "Tipo 4 (T0946-...)":
        pattern = re.compile(r"(T\d{4}-\d{4}-PE-[A-Z]{2}-[A-Z]{3}-\d{5}-\d{5}-[A-Z]-[A-Z]{3}-\d{4}-\d{2})\s+(.+?)(?=T\d{4}|$)", re.DOTALL)
    elif tipo == "Tipo 5 (Numerico puntato - 1.1.1 PDF)":
        pattern = re.compile(r"(\d+\.\d+\.\d+)\s+(.*?)\s+\d{2}/\d{2}/\d{4}\s+\d{2}\.\d{2}\s+[A-Z]+\s+[A-Z]+\s+\d{3}\s+[A-Z]?", re.DOTALL)
    elif tipo == "Tipo 6 (Codice T00EG00 complesso)":
        pattern = re.compile(r"(T00EG\d{2}[A-Z]{3}\d{2}[A-Z0-9])\s+(.*?)\s+A\d", re.DOTALL)
    else:
        return []

    matches = pattern.findall(testo_pdf)
    estratti = []
    for codice, descrizione in matches:
        estratti.append({
            "codice": codice.strip(),
            "descrizione": ' '.join(descrizione.strip().split()),
            "disciplina": "GEN",
            "formato": "A1",
            "scala": "1:100",
            "data": "n.d."
        })
    return estratti

def compila_excel(template_path, dati):
    wb = load_workbook(template_path)
    ws = wb["02_Elaborati"] if "02_Elaborati" in wb.sheetnames else wb.active

    for row in ws.iter_rows(min_row=7, max_row=ws.max_row, max_col=8):
        for cell in row:
            cell.value = None

    for i, row in enumerate(dati, start=7):
        ws[f"A{i}"] = row["codice"]
        ws[f"B{i}"] = row["descrizione"]
        ws[f"C{i}"] = row["disciplina"]
        ws[f"E{i}"] = row["formato"]
        ws[f"F{i}"] = row["scala"]
        ws[f"G{i}"] = row["data"]

    output_stream = BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    return output_stream

st.title("📁 Generatore ELENCO ELABORATI Excel da PDF (Definitivo)")

tipo_pdf = st.selectbox("Seleziona il tipo di PDF", [
    "Tipo 1 (PV014-PE-...)",
    "Tipo 2 (1.1.1)",
    "Tipo 3 (T00EG00...)",
    "Tipo 4 (T0946-...)",
    "Tipo 5 (Numerico puntato - 1.1.1 PDF)",
    "Tipo 6 (Codice T00EG00 complesso)"
])

uploaded_file = st.file_uploader("Carica il file PDF", type=["pdf"])

if uploaded_file:
    with st.spinner('Estrazione dati dal PDF...'):
        dati_estratti = estrai_dati(uploaded_file.getvalue(), tipo_pdf)

    if dati_estratti:
        st.success(f"✅ Estratti {len(dati_estratti)} elaborati.")

        if st.button("Genera Excel"):
            excel_stream = compila_excel("ELENCO ELABORATI.xlsx", dati_estratti)
            st.download_button(
                label="Scarica ELENCO ELABORATI.xlsx",
                data=excel_stream,
                file_name="ELENCO_ELABORATI_COMPILATO.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    else:
        st.error("⚠️ Nessun elaborato trovato nel PDF. Prova un altro tipo.")